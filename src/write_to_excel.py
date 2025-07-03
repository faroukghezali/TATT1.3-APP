import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


def write_to_excel(bon_number, bon_date):
    with sqlite3.connect("main.db") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT product_name,product_quantity,product_price,bus_id FROM article_bon WHERE bon_id = (SELECT id FROM bon WHERE bon_number = ?)",
            (bon_number,),
        )
        result = cursor.fetchall()

        default_font = Font(name="calibri", size=14, bold=True)
        wb = Workbook()
        # normal_style = wb.named_styles[0]
        # normal_style.font = default_font
        sheet = wb.active

        sheet.column_dimensions["A"].width = 75 / 7
        sheet.column_dimensions["B"].width = 511 / 7
        sheet.column_dimensions["C"].width = 85 / 7
        sheet.column_dimensions["D"].width = 163 / 7
        sheet.column_dimensions["E"].width = 215 / 7

        sheet.merge_cells("A3:E3")
        sheet.merge_cells("A4:E4")
        sheet.merge_cells("A11:B11")
        sheet.merge_cells("D11:E11")

        sheet["A3"] = "TATT SPA"
        sheet["A4"] = "TOURING ALGERIA TRANSPORT TRAVELS "
        sheet["A11"] = "DEPARTEMENT : MGX"
        sheet["D11"] = "DIRECTION: TATT SPA"

        sheet["A3"].font = Font(name="calibri", size=28, bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center")
        shhet["A"]

        sheet["E6"] = f"BIS-N°{bon_number}"
        sheet["E10"] = f"DATE : {bon_date}"
        row_index = 15
        for value_tuple in result:
            col_index = 2
            for value in value_tuple:
                sheet.cell(row_index, col_index, value=str(value))
                col_index += 1
            row_index += 1
        sheet.cell(row_index, 4, value=f"=SUM(D15:D{row_index - 1})")

        thin_style = Side(border_style="thin")
        table_border = Border(
            left=thin_style, bottom=thin_style, top=thin_style, right=thin_style
        )
        start_row = 15
        end_row = start_row + len(result)
        start_col = 1
        end_col = len(result[0]) + 1
        for row in sheet.iter_rows(
            min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
        ):
            for cell in row:
                cell.border = table_border
        wb.save(f"template/{bon_number}-test.xlsx")
        wb.close()

        # Check if form isnot empty
        # write into excel
