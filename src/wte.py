import os
import sqlite3
from datetime import datetime as dt
from datetime import timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook as lw
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from database import get_rapport_dactivity, get_total_by_category

# Constants for maintainability
HEADER_ROW = 3
TITLE_ROW = 4
INFO_ROWS = [6, 10, 11]
DATA_START_ROW = 15
COLUMN_WIDTHS = {
    "A": 75 / 7,
    "B": 511 / 7,
    "C": 85 / 7,
    "D": 163 / 7,
    "E": 215 / 7,
}
DEFAULT_FONT = Font(name="calibri", size=14, bold=True)
HEADER_FONT = Font(name="calibri", size=28, bold=True)


def write_to_excel(bon_number, bon_date):
    # Ensure template directory exists
    Path("template").mkdir(exist_ok=True)

    with sqlite3.connect("main.db") as conn:
        cursor = conn.cursor()
        # Improved query with JOIN
        cursor.execute(
            """SELECT a.product_name, a.product_quantity, a.product_price, a.bus_id 
               FROM article_bon a 
               JOIN bon b ON a.bon_id = b.id 
               WHERE b.bon_number = ?""",
            (bon_number,),
        )
        result = cursor.fetchall()

        cursor.execute(
            """SELECT SUM(a.product_price) 
               FROM article_bon a 
               JOIN bon b ON a.bon_id = b.id 
               WHERE b.bon_number = ?""",
            (bon_number,),
        )
        total = float(cursor.fetchone()[0])
        if not result:
            print("No data found for the given bon_number.")
            return

        wb = Workbook()
        sheet = wb.active

        # Set column widths
        for col, width in COLUMN_WIDTHS.items():
            sheet.column_dimensions[col].width = width

        # Merge and style headers
        sheet.merge_cells("A3:E3")
        sheet["A3"] = "TATT SPA"
        sheet["A3"].font = HEADER_FONT
        sheet["A3"].alignment = Alignment(horizontal="center")

        sheet.merge_cells("A4:E4")
        sheet["A4"] = "TOURING ALGERIA TRANSPORT TRAVELS"
        sheet["A4"].alignment = Alignment(horizontal="center")  # Centered A4
        sheet["A4"].font = Font(name="calibri", size=12, bold=True)

        # Additional headers (right-aligned)
        sheet["E6"] = f"BIS-N°{bon_number}"
        sheet["E6"].alignment = Alignment(horizontal="right")
        sheet["E10"] = f"DATE : {bon_date}"
        sheet["E10"].alignment = Alignment(horizontal="right")

        sheet.merge_cells("A11:B11")
        sheet["A11"] = "DEPARTEMENT : MGX"
        sheet.merge_cells("D11:E11")
        sheet["D11"] = "DIRECTION: TATT SPA"

        row_index = DATA_START_ROW
        header = ["id", "Produit", "Quantité", "Prix", "Matricule"]
        header_row = DATA_START_ROW - 1
        for col_idx, value in enumerate(header, start=1):
            sheet.cell(row=header_row, column=col_idx, value=str(value))

        for row_data in result:
            for col_idx, value in enumerate(row_data, start=2):
                if col_idx in [3, 4]:
                    try:
                        value = float(value)
                    except Exception:
                        pass
                sheet.cell(row=row_index, column=col_idx, value=value)
            row_index += 1

        if row_index > DATA_START_ROW:
            sum_cell = sheet.cell(
                row=row_index,
                column=4,
                value=total,
            )
            sum_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        end_data_row = row_index - 1 if result else DATA_START_ROW
        end_row = row_index if row_index > DATA_START_ROW else DATA_START_ROW
        thin_side = Side(border_style="thin")
        table_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        for row in sheet.iter_rows(
            min_row=header_row, max_row=end_row, min_col=1, max_col=5
        ):
            for cell in row:
                if cell.column in [3, 4]:
                    print(cell.value)
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                cell.border = table_border

        # Save and close
        # wb.save(f"template/{bon_number}-.xlsx")
        date = dt.now(timezone.utc).strftime("%Y-%m-%d")
        wb.save(f"{bon_number}-{date}.xlsx")
        wb.close()


def create_rapport(bon_number):
    green_fill = PatternFill(
        start_color="27ae60", end_color="2ecc71", fill_type="solid"
    )
    # template = "template/rapport.xlsx"
    # wb = lw(template)
    wb = Workbook()
    sheet = wb.active
    # Merge and style headers
    sheet.merge_cells("A3:H3")
    sheet["A3"] = "TATT SPA"
    sheet["A3"].font = HEADER_FONT
    sheet["A3"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A4:H4")
    sheet["A4"] = "ÉTAT DES TRAVAUX ET CONSOMMATIONS"
    sheet["A4"].alignment = Alignment(horizontal="center")  # Centered A4
    sheet["A4"].font = Font(name="calibri", size=12, bold=True)

    sheet.merge_cells("A6:H6")
    sheet["A6"] = "Direction des Activités Commerciales & Exploitation"
    sheet["A6"].alignment = Alignment(horizontal="center")  # Centered A4
    sheet["A6"].font = Font(name="calibri", size=14, bold=True)

    result = get_rapport_dactivity(bon_number)
    total = get_total_by_category(bon_number)

    header = [
        "N°TP",
        "MARQUE",
        "IMMATRICULATION",
        "Lubrifiant",
        "Ingrédient",
        "PneuMatique",
        "Pièce de Rechange",
        "TOTAL REPARATION",
    ]
    header_row = DATA_START_ROW - 1
    for col_idx, value in enumerate(header, start=1):
        sheet.cell(row=header_row, column=col_idx, value=str(value)).fill = green_fill

    row_index = DATA_START_ROW
    for row_data in result:
        for col_idx, value in enumerate(row_data, start=1):
            if col_idx > 3:
                try:
                    value = float(value)
                except Exception:
                    pass
            sheet.cell(row=row_index, column=col_idx, value=value)
        row_index += 1

    end_data_row = row_index - 1 if result else DATA_START_ROW
    end_row = row_index if row_index > DATA_START_ROW else DATA_START_ROW
    thin_side = Side(border_style="thin")
    table_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    x = 4

    for value in total:
        sheet.cell(row=end_row, column=x, value=f"{value:.2f}")
        x += 1
        if x == 8:
            sheet.cell(row=end_row, column=x, value=f"{sum(total):.2f}")

    for row in sheet.iter_rows(
        min_row=DATA_START_ROW, max_row=end_row, min_col=1, max_col=8
    ):
        for cell in row:
            if cell.column > 3:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            cell.border = table_border
    date = dt.now(timezone.utc).strftime("%Y-%m-%d")
    wb.save(f"rapport-{bon_number}-{date}.xlsx")
    wb.close()
